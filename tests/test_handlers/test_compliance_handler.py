from io import BytesIO
from statistics import mean
from types import SimpleNamespace
from typing import cast

from openpyxl import load_workbook
from xlsxwriter import Workbook

from handlers.reports.compliance_handler import (
    TOTAL_LOCATION,
    ComplianceReportXlsxWriter,
)
from helpers.constants import Cloud
from helpers.reports import Standard
from services.metadata import Metadata
from services.report_service import ReportService
from services.sharding import ShardPart, ShardsCollectionFactory


_STANDARD = Standard('Standard1', 'v1.0.0')
#: two controls, one rule is mapped to each of them
_FULL_COV = {_STANDARD: {'p1': 1, 'p2': 1}}


def _metadata(rules: dict[str, tuple[str, ...]]) -> Metadata:
    """
    Builds just enough of Metadata for coverage calculation: each given rule
    is mapped to the given controls of the single test standard
    """

    class _FakeMetadata(SimpleNamespace):
        def rule(self, name, /, **kwargs):
            controls = self.rules.get(name)
            if controls is None:
                return SimpleNamespace(standard={})
            return SimpleNamespace(
                standard={_STANDARD.name: {_STANDARD.version_str: controls}}
            )

        def domain(self, name, /):
            return SimpleNamespace(full_cov=_FULL_COV)

    return cast(Metadata, cast(object, _FakeMetadata(rules=rules)))


def test_total_over_whole_collection_is_not_the_average_of_regions():
    """
    Justifies why the endpoints calculate the total against the whole
    collection instead of aggregating the values already calculated per
    region: each region here covers only one control of the standard, so
    both are 50% covered whereas the account as a whole is fully covered
    """
    rs = ReportService(None, None)  # type: ignore[arg-type]
    metadata = _metadata({'rule-p1': ('p1',), 'rule-p2': ('p2',)})
    collection = ShardsCollectionFactory.from_cloud(Cloud.AWS)
    collection.put_parts(
        [
            ShardPart(policy='rule-p1', location='eu-west-1', resources=[]),
            ShardPart(policy='rule-p2', location='eu-central-1', resources=[]),
        ]
    )

    regions = {
        location: rs.calculate_coverages(
            successful=rs.get_standard_to_controls_to_rules(
                it=parts, metadata=metadata
            ),
            full=_FULL_COV,
        )
        for location, parts in rs.group_parts_iterator_by_location(
            rs.iter_successful_parts(collection)
        ).items()
    }
    total = rs.calculate_tenant_full_coverage(
        col=collection, metadata=metadata, cloud=Cloud.AWS
    )

    assert {
        location: data[_STANDARD] for location, data in regions.items()
    } == {'eu-west-1': 0.5, 'eu-central-1': 0.5}
    assert total[_STANDARD] == 1.0
    assert total[_STANDARD] != mean(
        data[_STANDARD] for data in regions.values()
    )


def test_xlsx_writes_total_as_the_last_row():
    coverages = {
        'eu-west-1': {_STANDARD.full_name: 0.5},
        'eu-central-1': {_STANDARD.full_name: 0.25},
        TOTAL_LOCATION: {_STANDARD.full_name: 0.75},
    }
    buffer = BytesIO()
    with Workbook(buffer) as workbook:
        ComplianceReportXlsxWriter(coverages).write(
            wb=workbook, wsh=workbook.add_worksheet('Compliance')
        )

    worksheet = load_workbook(BytesIO(buffer.getvalue()))['Compliance']

    assert worksheet['A1'].value == 'Regions'
    assert worksheet['B1'].value == _STANDARD.full_name
    assert [worksheet[f'A{row}'].value for row in (2, 3, 4)] == [
        'eu-west-1',
        'eu-central-1',
        TOTAL_LOCATION,
    ]
    # CellContent serializes everything but strings with json.dumps, so the
    # shared xlsx writer stores the coverages as text
    assert [worksheet[f'B{row}'].value for row in (2, 3, 4)] == [
        '0.5',
        '0.25',
        '0.75',
    ]
    assert worksheet.max_row == 4
