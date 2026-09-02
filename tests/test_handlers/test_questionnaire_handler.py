from io import BytesIO
from types import SimpleNamespace
from typing import cast

from openpyxl import load_workbook

from handlers.reports.questionnaire_handler import (
    ControlRow,
    QuestionnaireHandler,
    ReportScope,
    SheetLayout,
    Status,
    build_workbook,
    natural_key,
    short_job_id,
    slugify,
)
from helpers.constants import Cloud, Severity
from helpers.reports import Standard
from services.metadata import Metadata
from services.report_service import ReportService
from services.sharding import ShardPart, ShardsCollectionFactory


_STANDARD = Standard('CIS Kubernetes Benchmark', 'v1.7.0')


def _metadata(rules: dict) -> Metadata:
    class _FakeMetadata(SimpleNamespace):
        def rule(self, name, /, **kwargs):
            return self.rules.get(name, SimpleNamespace(standard={}))

    return cast(Metadata, cast(object, _FakeMetadata(rules=rules)))


def _rule(cloud, severity, controls):
    return SimpleNamespace(
        cloud=cloud,
        severity=severity,
        standard={_STANDARD.name: {_STANDARD.version_str: controls}},
    )


def _handler() -> QuestionnaireHandler:
    return QuestionnaireHandler(
        tenant_service=None,  # type: ignore[arg-type]
        job_service=None,  # type: ignore[arg-type]
        report_service=ReportService(None, None),  # type: ignore[arg-type]
        license_service=None,  # type: ignore[arg-type]
        platform_service=None,  # type: ignore[arg-type]
    )


def test_mapping_keeps_rules_and_selects_highest_cloud_severity():
    metadata = _metadata(
        {
            'rule-low': _rule(Cloud.AWS, Severity.LOW, ['1.1']),
            'rule-high': _rule(Cloud.AWS, Severity.HIGH, ['1.1']),
            'rule-other-cloud': _rule(Cloud.AZURE, Severity.HIGH, ['1.1']),
        }
    )

    result = QuestionnaireHandler.build_standard_control_to_rule_names(
        metadata=metadata, cloud=Cloud.AWS
    )

    assert result == {
        _STANDARD.name: {
            _STANDARD.version_str: {
                '1.1': {
                    'rules': ['rule-low', 'rule-high'],
                    'deprecated': [],
                    'severity': Severity.HIGH,
                },
            },
        },
    }


def test_mapping_sets_deprecated_rules_apart():
    metadata = _metadata(
        {
            'rule-low': _rule(Cloud.AWS, Severity.LOW, ['1.1']),
            'rule-high-deprecated': _rule(Cloud.AWS, Severity.HIGH, ['1.1']),
            # only looks for deprecated things, the suffix check must keep it
            'rule-deprecated-runtime': _rule(
                Cloud.AWS, Severity.MEDIUM, ['1.1']
            ),
        }
    )

    control = QuestionnaireHandler.build_standard_control_to_rule_names(
        metadata=metadata, cloud=Cloud.AWS
    )[_STANDARD.name][_STANDARD.version_str]['1.1']

    assert control['rules'] == ['rule-low', 'rule-deprecated-runtime']
    assert control['deprecated'] == ['rule-high-deprecated']
    # the severity of a deprecated rule must not leak into the control
    assert control['severity'] is Severity.MEDIUM


def test_resolve_standard_by_full_name():
    full_cov = {_STANDARD: {'1.1': 1}}
    assert (
        QuestionnaireHandler.resolve_standard(
            full_cov, 'cis kubernetes benchmark v1.7.0'
        )
        is _STANDARD
    )
    assert QuestionnaireHandler.resolve_standard(full_cov, 'nope') is None


def test_build_rows_aggregates_regions_with_failure_precedence():
    collection = ShardsCollectionFactory.from_cloud(Cloud.AWS)
    collection.put_parts(
        [
            ShardPart(policy='rule-a', location='eu-west-1', resources=[]),
            ShardPart(policy='rule-b', location='eu-west-1', resources=[{}]),
            ShardPart(policy='rule-c', location='eu-central-1', resources=[]),
        ]
    )
    metadata = _metadata(
        {
            'rule-a': _rule(Cloud.AWS, Severity.LOW, ['1.1']),
            'rule-b': _rule(Cloud.AWS, Severity.HIGH, ['1.1']),
            'rule-c': _rule(Cloud.AWS, Severity.MEDIUM, ['1.1']),
            'rule-d': _rule(Cloud.AWS, Severity.MEDIUM, ['1.2']),
        }
    )

    rows = _handler().build_rows(
        collection=collection,
        metadata=metadata,
        cloud=Cloud.AWS,
        standard=_STANDARD,
        full_controls={'1.1': 3, '1.2': 1},
    )

    assert [row.control_id for row in rows] == ['1.1', '1.2']
    first, second = rows
    # rule-b failed in eu-west-1 therefore it is failed everywhere
    assert first.failed_rules == ['rule-b']
    assert first.successful_rules == ['rule-a', 'rule-c']
    assert first.status is Status.FAIL
    assert first.severity == Severity.HIGH.value
    # rule-a and rule-c succeeded in different regions but both count for
    # account-level coverage.
    assert first.coverage == 2 / 3

    # nothing was scanned for 1.2 -> not evaluated
    assert second.not_evaluated_rules == ['rule-d']
    assert second.status is Status.NOT_EVALUATED
    assert second.coverage == 0.0


def test_build_rows_drops_deprecated_rules_from_status_and_rule_ids():
    collection = ShardsCollectionFactory.from_cloud(Cloud.AWS)
    collection.put_parts(
        [
            ShardPart(policy='rule-a', location='eu-west-1', resources=[]),
            # deprecated and failing: must not turn the control into Fail
            ShardPart(
                policy='rule-b-deprecated',
                location='eu-west-1',
                resources=[{}],
            ),
        ]
    )
    metadata = _metadata(
        {
            'rule-a': _rule(Cloud.AWS, Severity.LOW, ['1.1']),
            'rule-b-deprecated': _rule(Cloud.AWS, Severity.HIGH, ['1.1']),
        }
    )

    # full_cov comes from LM and counts the deprecated rule as well
    (row,) = _handler().build_rows(
        collection=collection,
        metadata=metadata,
        cloud=Cloud.AWS,
        standard=_STANDARD,
        full_controls={'1.1': 2},
    )

    assert row.successful_rules == ['rule-a']
    assert row.failed_rules == []
    assert row.not_evaluated_rules == []
    assert row.status is Status.PASS
    assert row.rule_short_ids == 'rule-a'
    # the severity of the deprecated rule is not reported either
    assert row.severity == Severity.LOW.value
    # Coverage still uses the unmodified `full_cov` denominator, so it does not
    # reach 100% even though every reported rule passed.
    assert row.coverage == 0.5


def test_natural_key_sorts_controls_naturally():
    assert sorted(['1.10', '1.2', '1.1'], key=natural_key) == [
        '1.1',
        '1.2',
        '1.10',
    ]


def test_layout_omits_hierarchy_columns_when_not_provided():
    rows = [ControlRow(control_id='1.1')]
    layout = SheetLayout.for_rows(rows)

    assert layout.headers == [
        'Control',
        'Severity',
        'Status',
        'Coverage',
        'Rule IDs',
    ]
    assert layout.row_values(rows[0]) == [
        '1.1',
        None,
        Status.NOT_EVALUATED.value,
        0.0,
        None,
    ]


def test_layout_supports_injected_hierarchy():
    rows = [
        ControlRow(
            control_id='1.1',
            section_id='1',
            section_title='Section one',
            sub_sections=[('1.a', 'Sub one')],
            control_title='Do the thing',
        )
    ]
    layout = SheetLayout.for_rows(rows)

    assert layout.headers == [
        'Section',
        'Title',
        'Sub-section',
        'Sub-title',
        'Control',
        'Severity',
        'Status',
        'Control Title',
        'Coverage',
        'Rule IDs',
    ]


def test_build_workbook_contains_all_sheets():
    rows = [
        ControlRow(
            control_id='1.1',
            severity=Severity.HIGH.value,
            coverage=1.0,
            successful_rules=['ecc-aws-001-something'],
        )
    ]
    buffer = BytesIO()
    build_workbook(rows, _STANDARD, ReportScope.JOB).save(buffer)
    buffer.seek(0)

    wb = load_workbook(buffer)
    assert wb.sheetnames == [
        'Introduction',
        'Compliance',
        'Summary',
        '_ChartData',
    ]
    ws = wb['Compliance']
    assert [cell.value for cell in ws[1]] == [
        'Control',
        'Severity',
        'Status',
        'Coverage',
        'Rule IDs',
    ]
    assert [cell.value for cell in ws[2]] == [
        '1.1',
        'High',
        'Pass',
        1.0,
        '001',
    ]


def _intro_texts(scope: ReportScope) -> list[str]:
    buffer = BytesIO()
    build_workbook([ControlRow(control_id='1.1')], _STANDARD, scope).save(
        buffer
    )
    buffer.seek(0)
    ws = load_workbook(buffer)['Introduction']
    return [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]


def test_introduction_describes_job_scope():
    texts = _intro_texts(ReportScope.JOB)

    assert 'I. Scope' in texts
    assert 'Report type: Job questionnaire.' in texts
    assert any('one single scan job' in text for text in texts)
    assert any(
        'jobs-<standard>-<job id>-questionnaire.xlsx' in text
        and 'last segment of the job UUID' in text
        for text in texts
    )
    # sections are renumbered once Scope is inserted
    assert 'II. Structure' in texts
    assert 'III. Description' in texts
    assert 'Notice' in texts


def test_introduction_describes_accumulated_scope():
    texts = _intro_texts(ReportScope.ACCUMULATED)

    assert 'Report type: Accumulated (tenant) questionnaire.' in texts
    assert any('latest accumulated state of the tenant' in t for t in texts)
    assert any(
        'accumulated-<standard>-<tenant name>-questionnaire.xlsx' in text
        for text in texts
    )


def test_slugify_standard_full_name():
    assert slugify('CIS Kubernetes Benchmark v1.7.0') == (
        'cis_kubernetes_benchmark_v1_7_0'
    )
    assert slugify('CIS Controls') == 'cis_controls'


def test_short_job_id_takes_last_uuid_segment():
    assert short_job_id('20d3a4e1-6a0f-4c0e-9c11-1f2d3e4f5a6b') == '1f2d3e4f5a6b'
    assert short_job_id('plain') == 'plain'


def test_filenames_follow_the_agreed_format():
    assert ReportScope.JOB.filename(_STANDARD, '1f2d3e4f5a6b') == (
        'jobs-cis_kubernetes_benchmark_v1_7_0-1f2d3e4f5a6b-questionnaire.xlsx'
    )
    assert ReportScope.ACCUMULATED.filename(_STANDARD, 'MY-TENANT') == (
        'accumulated-cis_kubernetes_benchmark_v1_7_0-MY-TENANT'
        '-questionnaire.xlsx'
    )



