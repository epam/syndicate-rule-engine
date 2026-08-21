"""
Standard Questionnaire Generator
================================
Builds an Excel workbook that maps every control of a security standard to its
audit status, coverage and the rules that evaluate it.

Sheets
------
  1. '<Standard> <version>' - one row per control (Excel table + conditional
     formatting, so manual edits stay color-coded).
  2. 'Summary'              - formula-driven statistics table + two pie charts.
  3. 'Checks'              - formula-driven statistics + pie and stacked bars.
  ('_ChartData' is a hidden helper sheet feeding the bar-of-pie chart.)

Data sources
------------
  Compliance report (per scan) : coverage, severity, rule pass/fail lists.
  standards_schema (static)    : section / sub-section / control hierarchy.

Multi-region aggregation
------------------------
  failed        = union over all regions (worst case wins)
  successful    = union minus failed
  not_evaluated = union minus failed and successful

Status logic
------------
  Fail          - at least one rule failed
  Pass          - every rule succeeded
  Not Evaluated - no failure but at least one rule unevaluated (or no rules)

Coverage is never recalculated; the report's ``total`` value is used verbatim.

Usage
-----
    python questionnaire.py \
        --standard   'CIS Controls v7' \
        --compliance job-compliance.json \
        --schemas    mappings \
        --output     cis_questionnaire.xlsx

Note: schema files are Python modules and are *executed* on load; only point
``--schemas`` at a trusted repository.
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import runpy
import sys
from collections import Counter
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass, field
from enum import Enum
from itertools import chain
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, ProjectedPieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger('questionnaire')


class QuestionnaireError(RuntimeError):
    """Any error that should end the run with a friendly CLI message."""


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

HEX_HEADER_BG = '4472C4'
HEX_HEADER_FONT = 'FFFFFF'
HEX_GREEN_PASTEL = 'C6EFCE'
HEX_GREEN_FONT = '006100'
HEX_RED_PASTEL = 'FFC7CE'
HEX_RED_FONT = '9C0006'
HEX_GRAY = 'BFBFBF'
HEX_GRAY_FONT = '404040'
HEX_BORDER_BLUE = '9DC3E6'

CHART_FAIL_COLOR = 'ED7D31'
CHART_PASS_COLOR = '70AD47'
CHART_NA_COLOR = 'A5A5A5'

#: Severities that the compliance report actually uses in charts/tables.
REPORT_SEVERITIES: tuple[str, ...] = ('High', 'Medium', 'Low')
SEVERITY_CHART_COLORS = {'High': 'FF3300', 'Medium': 'FFC000', 'Low': 'BDD7EE'}
SEVERITY_LABEL_FONTS = {'High': 'FFFFFF', 'Medium': '000000', 'Low': '000000'}
#: Conditional-format fills for the Severity column of the questionnaire.
SEVERITY_CELL_FILLS = {
    'High': 'FF3300',
    'Medium': 'FFCC00',
    'Low': 'FFFF99',
    'Info': 'A6C9EC',
}

BORDER_SIDE = Side(style='thin', color=HEX_BORDER_BLUE)
BLUE_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
_THIN = Side(style='thin')
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WHITE_FILL = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')

ALIGN_LEFT = Alignment(horizontal='left', vertical='top')
ALIGN_CENTER = Alignment(horizontal='center', vertical='top')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)

PERCENT_FORMAT = '0%'
HEADER_ROW = 1
FIRST_DATA_ROW = 2
MAX_SHEET_TITLE = 31
INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')

TABLE_NAME = 'StandardQuestionnaire'
CHART_DATA_SHEET = '_ChartData'
CHART_FONT = 'Aptos Narrow'
CHART_TITLE_SIZE = 1600
CHART_LEGEND_SIZE = 900


# ─────────────────────────────────────────────────────────────────────────────
# Domain model
# ─────────────────────────────────────────────────────────────────────────────

class Status(str, Enum):
    PASS = 'Pass'
    FAIL = 'Fail'
    NOT_EVALUATED = 'Not Evaluated'


STATUS_FILL_COLORS = {
    Status.PASS.value: HEX_GREEN_PASTEL,
    Status.FAIL.value: HEX_RED_PASTEL,
    Status.NOT_EVALUATED.value: HEX_GRAY,
}
STATUS_FONT_COLORS = {
    Status.PASS.value: HEX_GREEN_FONT,
    Status.FAIL.value: HEX_RED_FONT,
    Status.NOT_EVALUATED.value: HEX_GRAY_FONT,
}


def short_rule_id(rule_id: str) -> str:
    """'ecc-aws-001-some_name' -> '001'; anything else is returned unchanged."""
    parts = rule_id.split('-')
    return parts[2] if len(parts) >= 3 and parts[2].isdigit() else rule_id


@dataclass
class ControlRow:
    """One control of the standard, enriched with compliance results."""

    section_id: str
    section_title: str
    control_id: str
    control_title: str = ''
    #: (id, title) pairs between the top-level section and the control.
    sub_sections: list[tuple[str, str]] = field(default_factory=list)
    coverage: float = 0.0
    severity: str = ''
    successful_rules: list[str] = field(default_factory=list)
    failed_rules: list[str] = field(default_factory=list)
    not_evaluated_rules: list[str] = field(default_factory=list)

    @property
    def status(self) -> Status:
        if self.failed_rules:
            return Status.FAIL
        if self.not_evaluated_rules or not self.successful_rules:
            return Status.NOT_EVALUATED
        return Status.PASS

    @property
    def rule_short_ids(self) -> str:
        ids = {
            short_rule_id(rule)
            for rule in chain(self.successful_rules, self.failed_rules, self.not_evaluated_rules)
        }
        return ', '.join(sorted(ids))


@dataclass
class ControlCompliance:
    """Compliance results for one control, aggregated over all regions."""

    coverage: float | None = None
    severity: str | None = None
    successful: set[str] = field(default_factory=set)
    failed: set[str] = field(default_factory=set)
    not_evaluated: set[str] = field(default_factory=set)

    def absorb(self, control: Mapping[str, Any]) -> None:
        """Merge one region's entry for this control (first values win)."""
        if self.coverage is None:
            self.coverage = _as_float(control.get('total'))
        if self.severity is None and control.get('severity') is not None:
            self.severity = str(control['severity'])

        rules = control.get('rules')
        if not isinstance(rules, Mapping):
            return
        self.failed |= set(rules.get('failed') or ())
        self.successful |= set(rules.get('successful') or ())
        self.not_evaluated |= set(rules.get('not_evaluated') or ())

    def resolve_overlaps(self) -> None:
        """A rule that failed anywhere counts as failed; then as successful."""
        self.successful -= self.failed
        self.not_evaluated -= self.failed | self.successful


@dataclass(frozen=True)
class Statistics:
    total: int
    passed: int
    failed: int
    not_evaluated: int
    applicable_coverage: float
    total_coverage: float

    @classmethod
    def from_rows(cls, rows: Sequence[ControlRow]) -> 'Statistics':
        counts = Counter(row.status for row in rows)
        applicable = [row.coverage for row in rows if row.status is not Status.NOT_EVALUATED]
        return cls(
            total=len(rows),
            passed=counts[Status.PASS],
            failed=counts[Status.FAIL],
            not_evaluated=counts[Status.NOT_EVALUATED],
            applicable_coverage=_mean(applicable),
            total_coverage=_mean([row.coverage for row in rows]),
        )

    def render(self, output_path: Path) -> str:
        return (
            f'✓ Saved: {output_path}  ({self.total} controls)\n'
            f'  Pass          : {self.passed}\n'
            f'  Fail          : {self.failed}\n'
            f'  Not Evaluated : {self.not_evaluated}\n'
            f'  SRE Applicable Coverage : {self.applicable_coverage:.0%}\n'
            f'  Total Coverage          : {self.total_coverage:.0%}'
        )


def _as_float(value: Any) -> float | None:
    try:
        return float(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def _mean(values: Sequence[float]) -> float:
    return sum(values) / len(values) if values else 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Schema loading
# ─────────────────────────────────────────────────────────────────────────────

def parse_standard_name(name: str) -> tuple[str, str | None]:
    """'CIS Controls v7' -> ('CIS Controls', 'v7'); 'AWS Config' -> (..., None)."""
    match = re.search(r'\s+(v[\d.]+)$', name, re.IGNORECASE)
    if match:
        return name[: match.start()].strip(), match.group(1)
    return name.strip(), None


def find_schema_file(schemas_root: Path, base_name: str) -> Path:
    """Locate ``<schemas_root>/standards_schema/<Base_Name>.py``."""
    schema_dir = schemas_root / 'standards_schema'
    if not schema_dir.is_dir():
        raise QuestionnaireError(f'missing schema directory: {schema_dir}')

    wanted = f'{base_name.replace(' ', '_')}.py'
    for candidate in (schema_dir / wanted, *schema_dir.glob('*.py')):
        if candidate.name.lower() == wanted.lower() and candidate.is_file():
            return candidate

    available = sorted(p.stem for p in schema_dir.glob('*.py') if p.stem != '__init__')
    raise QuestionnaireError(
        f'no schema file found for `{base_name}` in {schema_dir}. '
        f'Available schemas: {", ".join(available)}'
    )


def load_version_block(schema_file: Path, version: str | None) -> tuple[dict, str]:
    """Execute the schema module and return (version block, resolved version)."""
    namespace = runpy.run_path(str(schema_file))
    data = namespace.get('full') or namespace.get('tech') or {}
    if not data:
        raise QuestionnaireError(f'{schema_file} defines neither `full` nor `tech`')

    if version is None:
        resolved = next(iter(data))
    else:
        resolved = next((key for key in data if key.lower() == version.lower()), None)
        if resolved is None:
            raise QuestionnaireError(
                f'version `{version}` not found in {schema_file.name}. '
                f'Available: {list(data)}'
            )
    return data[resolved], resolved


def parse_schema(version_block: Mapping[str, Any]) -> list[ControlRow]:
    """Flatten a version block into ControlRows (schema data only)."""
    rows: list[ControlRow] = []
    for section_id, node in (version_block.get('sections') or {}).items():
        node = node or {}
        rows.extend(_iter_controls(node, [(section_id, node.get('section', ''))]))
    return rows


def _iter_controls(
    node: Mapping[str, Any],
    ancestors: list[tuple[str, str]],
) -> Iterator[ControlRow]:
    """Depth-first walk yielding one ControlRow per control below *node*."""
    (section_id, section_title), *sub_sections = ancestors
    for control_id, control in (node.get('controls') or {}).items():
        yield ControlRow(
            section_id=section_id,
            section_title=section_title,
            sub_sections=list(sub_sections),
            control_id=control_id,
            control_title=(control or {}).get('control_title', ''),
        )
    for sub_id, sub_node in (node.get('sections') or {}).items():
        sub_node = sub_node or {}
        yield from _iter_controls(sub_node, [*ancestors, (sub_id, sub_node.get('section', ''))])


# ─────────────────────────────────────────────────────────────────────────────
# Compliance report parsing
# ─────────────────────────────────────────────────────────────────────────────

def parse_compliance(report: Mapping[str, Any], standard_name: str) -> dict[str, ControlCompliance]:
    """Aggregate ``{region: {standard: {controls: ...}}}`` for one standard."""
    aggregated: dict[str, ControlCompliance] = {}
    for region in report.values():
        standard = _standard_block(region, standard_name)
        if standard is None:
            continue
        for control_id, control in (standard.get('controls') or {}).items():
            aggregated.setdefault(control_id, ControlCompliance()).absorb(control or {})

    for compliance in aggregated.values():
        compliance.resolve_overlaps()
    return aggregated


def _standard_block(region: Any, standard_name: str) -> Mapping[str, Any] | None:
    if not isinstance(region, Mapping):
        return None
    wanted = standard_name.lower()
    return next((v for k, v in region.items() if k.lower() == wanted), None)


def apply_compliance(rows: Sequence[ControlRow], compliance: Mapping[str, ControlCompliance]) -> None:
    """Copy compliance results onto the matching ControlRows (in place)."""
    for row in rows:
        result = compliance.get(row.control_id)
        if result is None:
            continue
        row.coverage = result.coverage or 0.0
        row.severity = result.severity or ''
        row.successful_rules = sorted(result.successful)
        row.failed_rules = sorted(result.failed)
        row.not_evaluated_rules = sorted(result.not_evaluated)


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers - styles, formulas, charts
# ─────────────────────────────────────────────────────────────────────────────

def fill(color: str) -> PatternFill:
    return PatternFill('solid', start_color=color, end_color=color)


def sheet_title(name: str) -> str:
    """Excel rejects >31 chars and []:*?/\\ in sheet names."""
    return INVALID_SHEET_CHARS.sub('-', name)[:MAX_SHEET_TITLE].strip() or 'Sheet1'


def table_ref(column: str) -> str:
    return f'{TABLE_NAME}[{column}]'


STATUS_REF = table_ref('Status')
SEVERITY_REF = table_ref('Severity')
COVERAGE_REF = table_ref('Coverage')


def count_status(status: Status) -> str:
    return f'=COUNTIF({STATUS_REF},"{status.value}")'


def count_status_severity(status: Status, severity: str) -> str:
    return f'=COUNTIFS({STATUS_REF},"{status.value}",{SEVERITY_REF},"{severity}")'


def stat_cell(
    ws: Worksheet,
    row: int,
    column: int,
    value: Any = None,
    *,
    bg: str | None = None,
    font: Font | None = None,
    align: Alignment = ALIGN_CENTER,
):
    """Bordered cell used by the statistics tables of Summary/Checks."""
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = THIN_BORDER
    cell.alignment = align
    if bg:
        cell.fill = fill(bg)
    if font:
        cell.font = font
    return cell


def write_sheet_title(ws: Worksheet, text: str) -> None:
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, size=14)


def set_column_widths(ws: Worksheet, widths: Mapping[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def autofit_columns(ws: Worksheet, *, max_row: int, limit: int = 50, minimum: int = 10) -> None:
    for column in ws.iter_cols(min_row=1, max_row=max_row):
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        letter = get_column_letter(column[0].column)
        ws.column_dimensions[letter].width = max(minimum, min(longest + 2, limit))


def add_text_conditional_format(
    ws: Worksheet,
    column: int,
    last_row: int,
    fills: Mapping[str, str],
    fonts: Mapping[str, str] | None = None,
) -> None:
    """Color cells by their own text, so manual edits stay color-coded."""
    if last_row < FIRST_DATA_ROW:
        return
    letter = get_column_letter(column)
    cell_range = f'{letter}{FIRST_DATA_ROW}:{letter}{last_row}'
    for text, bg in fills.items():
        font_color = (fonts or {}).get(text)
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'UPPER(${letter}{FIRST_DATA_ROW})="{text.upper()}"'],
                fill=fill(bg),
                font=Font(color=font_color) if font_color else None,
            ),
        )


def _chart_text(size: int) -> RichText:
    return RichText(
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=size, latin=DrawingFont(typeface=CHART_FONT))
                )
            )
        ]
    )


def style_chart(chart, *, title: str, width: float = 10, height: float = 6) -> None:
    chart.title = title
    chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.roundedCorners = False
    if chart.title is not None:
        chart.title.overlay = False
        chart.title.txPr = _chart_text(CHART_TITLE_SIZE)
    chart.legend.position = 'r'
    chart.legend.overlay = False
    chart.legend.txPr = _chart_text(CHART_LEGEND_SIZE)
    chart.width = width
    chart.height = height


def set_data_labels(
    chart,
    *,
    show_value: bool = False,
    show_percent: bool = False,
    position: str | None = None,
) -> None:
    labels = DataLabelList()
    labels.showVal = show_value
    labels.showPercent = show_percent
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    if position:
        labels.dLblPos = position
    chart.dataLabels = labels


def color_points(series, colors: Sequence[str]) -> None:
    """Color individual pie slices, in data order."""
    for idx, color in enumerate(colors):
        point = DataPoint(idx=idx)
        point.graphicalProperties = GraphicalProperties(solidFill=color)
        series.data_points.append(point)


def color_series(series, color: str) -> None:
    series.graphicalProperties = GraphicalProperties(solidFill=color)


def style_value_axes(chart) -> None:
    for axis in (chart.x_axis, chart.y_axis):
        axis.delete = False
        axis.txPr = _chart_text(CHART_LEGEND_SIZE)
    chart.y_axis.tickLblPos = 'low'


# ─────────────────────────────────────────────────────────────────────────────
# Questionnaire sheet
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class SheetLayout:
    """Column geometry of the questionnaire sheet (1-based indices)."""

    sub_depth: int

    @classmethod
    def for_rows(cls, rows: Sequence[ControlRow]) -> 'SheetLayout':
        return cls(max((len(row.sub_sections) for row in rows), default=0))

    @property
    def sub_headers(self) -> list[str]:
        if self.sub_depth == 0:
            return []
        if self.sub_depth == 1:
            return ['Sub-section', 'Sub-title']
        return [
            header
            for level in range(2, self.sub_depth + 2)
            for header in (f'Sub-section (l{level})', f'Sub-title (l{level})')
        ]

    @property
    def headers(self) -> list[str]:
        return [
            'Section',
            'Title',
            *self.sub_headers,
            'Control',
            'Severity',
            'Status',
            'Control Title',
            'Coverage',
            'Rule IDs',
        ]

    @property
    def width(self) -> int:
        return len(self.headers)

    @property
    def control(self) -> int:
        return 3 + 2 * self.sub_depth

    @property
    def severity(self) -> int:
        return self.control + 1

    @property
    def status(self) -> int:
        return self.control + 2

    @property
    def control_title(self) -> int:
        return self.control + 3

    @property
    def coverage(self) -> int:
        return self.control + 4

    def alignment(self, column: int) -> Alignment:
        if column == self.status:
            return ALIGN_CENTER
        if column == self.control_title:
            return ALIGN_WRAP
        return ALIGN_LEFT


def _row_values(row: ControlRow, layout: SheetLayout) -> list[Any]:
    """Cell values for one control, always ``layout.width`` long."""
    values: list[Any] = [row.section_id, row.section_title]
    for level in range(layout.sub_depth):
        pair = row.sub_sections[level] if level < len(row.sub_sections) else (None, None)
        values.extend(pair)
    values.extend(
        [
            row.control_id,
            row.severity or None,
            row.status.value,
            row.control_title,
            row.coverage,
            row.rule_short_ids or None,
        ]
    )
    return values


def build_questionnaire_sheet(ws: Worksheet, rows: Sequence[ControlRow], layout: SheetLayout) -> None:
    for column, header in enumerate(layout.headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=column, value=header)
        cell.fill = fill(HEX_HEADER_BG)
        cell.font = Font(color=HEX_HEADER_FONT, bold=True)
        cell.alignment = ALIGN_HEADER
        cell.border = BLUE_BORDER
    ws.row_dimensions[HEADER_ROW].height = 30

    for offset, row in enumerate(rows):
        excel_row = FIRST_DATA_ROW + offset
        for column, value in enumerate(_row_values(row, layout), start=1):
            cell = ws.cell(row=excel_row, column=column, value=value)
            cell.alignment = layout.alignment(column)
            cell.border = BLUE_BORDER
            cell.fill = WHITE_FILL
            if column == layout.coverage:
                cell.number_format = PERCENT_FORMAT

    last_row = FIRST_DATA_ROW + len(rows) - 1
    _add_questionnaire_table(ws, layout, last_row)
    add_text_conditional_format(ws, layout.severity, last_row, SEVERITY_CELL_FILLS)
    add_text_conditional_format(ws, layout.status, last_row, STATUS_FILL_COLORS, STATUS_FONT_COLORS)
    _add_coverage_totals(ws, layout, last_row)

    autofit_columns(ws, max_row=last_row)
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1).coordinate


def _add_questionnaire_table(ws: Worksheet, layout: SheetLayout, last_row: int) -> None:
    ref = f'A{HEADER_ROW}:{get_column_letter(layout.width)}{last_row}'
    table = Table(displayName=TABLE_NAME, ref=ref)
    table.autoFilter = AutoFilter(ref=ref)
    table.tableColumns = [
        TableColumn(id=index, name=header) for index, header in enumerate(layout.headers, start=1)
    ]
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_coverage_totals(ws: Worksheet, layout: SheetLayout, last_row: int) -> None:
    """Coverage averages written *outside* the table so they are not filtered."""
    totals = (
        (
            'SRE Applicable Coverage',
            f'=IFERROR(AVERAGEIF({STATUS_REF},"<>{Status.NOT_EVALUATED.value}",{COVERAGE_REF}),0)',
        ),
        ('Total Coverage', f'=IFERROR(AVERAGE({COVERAGE_REF}),0)'),
    )
    for offset, (label, formula) in enumerate(totals, start=2):
        excel_row = last_row + offset
        label_cell = ws.cell(row=excel_row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = ALIGN_LEFT

        value_cell = ws.cell(row=excel_row, column=layout.coverage, value=formula)
        value_cell.font = Font(bold=True)
        value_cell.alignment = ALIGN_LEFT
        value_cell.number_format = PERCENT_FORMAT


# ─────────────────────────────────────────────────────────────────────────────
# Summary sheet
# ─────────────────────────────────────────────────────────────────────────────

SUMMARY_TABLE_ROW = 15
#: (label, status, fill, font color, split by severity)
SUMMARY_ROWS: tuple[tuple[str, Status, str, str, bool], ...] = (
    ('Fail', Status.FAIL, HEX_RED_PASTEL, HEX_RED_FONT, True),
    ('Pass', Status.PASS, HEX_GREEN_PASTEL, HEX_GREEN_FONT, True),
    ('Not Applicable', Status.NOT_EVALUATED, HEX_GRAY, HEX_GRAY_FONT, False),
)


def build_summary_sheet(wb: Workbook, standard_name: str, version: str) -> None:
    """Formula-driven statistics table plus two pie charts."""
    ws = wb.create_sheet('Summary')
    helper = wb.create_sheet(CHART_DATA_SHEET)
    helper.sheet_state = 'hidden'

    write_sheet_title(ws, f'Summary by {standard_name} {version}')
    _write_summary_table(ws)
    _write_bar_of_pie_data(helper)

    ws.add_chart(_total_coverage_pie(ws), 'A2')
    ws.add_chart(_applicable_coverage_pie(helper), 'F2')

    set_column_widths(ws, {'A': 18, 'B': 12, 'C': 10, 'D': 10, 'E': 10, 'F': 3})


def _write_summary_table(ws: Worksheet) -> None:
    stat_cell(ws, SUMMARY_TABLE_ROW, 1, font=Font(bold=True))
    stat_cell(ws, SUMMARY_TABLE_ROW, 2, 'Controls', font=Font(bold=True))
    for offset, severity in enumerate(REPORT_SEVERITIES):
        stat_cell(
            ws,
            SUMMARY_TABLE_ROW,
            3 + offset,
            severity,
            bg=SEVERITY_CHART_COLORS[severity],
            font=Font(color=SEVERITY_LABEL_FONTS[severity], bold=True),
        )

    for index, (label, status, bg, font_color, by_severity) in enumerate(SUMMARY_ROWS, start=1):
        excel_row = SUMMARY_TABLE_ROW + index
        stat_cell(ws, excel_row, 1, label, bg=bg, font=Font(color=font_color, bold=True))
        stat_cell(ws, excel_row, 2, count_status(status))
        for offset, severity in enumerate(REPORT_SEVERITIES):
            value = count_status_severity(status, severity) if by_severity else '-'
            stat_cell(ws, excel_row, 3 + offset, value)

    total_row = SUMMARY_TABLE_ROW + len(SUMMARY_ROWS) + 1
    stat_cell(ws, total_row, 1, 'Total', font=Font(bold=True))
    stat_cell(ws, total_row, 2, f'=ROWS({STATUS_REF})', font=Font(bold=True))
    for offset in range(len(REPORT_SEVERITIES)):
        stat_cell(ws, total_row, 3 + offset)


def _write_bar_of_pie_data(helper: Worksheet) -> None:
    entries = [('Pass', count_status(Status.PASS))] + [
        (severity, count_status_severity(Status.FAIL, severity)) for severity in REPORT_SEVERITIES
    ]
    for index, (label, formula) in enumerate(entries, start=1):
        helper.cell(row=index, column=1, value=label)
        helper.cell(row=index, column=2, value=formula)


def _total_coverage_pie(ws: Worksheet) -> PieChart:
    first, last = SUMMARY_TABLE_ROW + 1, SUMMARY_TABLE_ROW + len(SUMMARY_ROWS)
    chart = PieChart()
    chart.add_data(Reference(ws, min_col=2, min_row=first, max_row=last), titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    color_points(chart.series[0], (CHART_FAIL_COLOR, CHART_PASS_COLOR, CHART_NA_COLOR))
    set_data_labels(chart, show_percent=True)
    style_chart(chart, title='Total Coverage')
    return chart


def _applicable_coverage_pie(helper: Worksheet) -> ProjectedPieChart:
    last = 1 + len(REPORT_SEVERITIES)
    chart = ProjectedPieChart()
    chart.type = 'bar'
    chart.splitType = 'pos'
    chart.splitPos = len(REPORT_SEVERITIES)
    chart.add_data(Reference(helper, min_col=2, min_row=1, max_row=last), titles_from_data=False)
    chart.set_categories(Reference(helper, min_col=1, min_row=1, max_row=last))

    series = chart.series[0]
    color_series(series, CHART_FAIL_COLOR)  # combined 'other' slice of the main pie
    color_points(
        series,
        (CHART_PASS_COLOR, *(SEVERITY_CHART_COLORS[s] for s in REPORT_SEVERITIES)),
    )
    set_data_labels(chart, show_percent=True)
    style_chart(chart, title='Applicable Coverage')
    return chart


# ─────────────────────────────────────────────────────────────────────────────
# Checks sheet
# ─────────────────────────────────────────────────────────────────────────────

CHECKS_HEADER_ROW = 9
#: (label, status, fill, font color) - one data row per status
CHECKS_ROWS: tuple[tuple[str, Status, str, str], ...] = (
    ('Fail', Status.FAIL, HEX_RED_PASTEL, HEX_RED_FONT),
    ('Pass', Status.PASS, HEX_GREEN_PASTEL, HEX_GREEN_FONT),
)
CHECKS_FIRST_DATA_ROW = CHECKS_HEADER_ROW + 1
CHECKS_LAST_DATA_ROW = CHECKS_HEADER_ROW + len(CHECKS_ROWS)
CHECKS_LAST_COL = 1 + len(REPORT_SEVERITIES)


def build_checks_sheet(wb: Workbook) -> None:
    """Summary stats, a failed-by-severity pie and two stacked bar charts."""
    ws = wb.create_sheet('Checks')
    write_sheet_title(ws, 'Summary by Checks')

    stats = (
        ('Total checks:', f'=ROWS({STATUS_REF})'),
        ('Failed:', count_status(Status.FAIL)),
        ('Passed:', count_status(Status.PASS)),
    )
    for excel_row, (label, formula) in enumerate(stats, start=4):
        ws.cell(row=excel_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=excel_row, column=2, value=formula)

    _write_checks_table(ws)
    ws.add_chart(_failed_by_severity_pie(ws), 'G3')
    ws.add_chart(_severity_vs_status_bar(ws), 'A14')
    ws.add_chart(_status_vs_severity_bar(ws), 'A24')

    set_column_widths(ws, {'A': 16, 'B': 12, 'C': 12, 'D': 12})


def _write_checks_table(ws: Worksheet) -> None:
    stat_cell(ws, CHECKS_HEADER_ROW, 1)
    for offset, severity in enumerate(REPORT_SEVERITIES):
        stat_cell(
            ws,
            CHECKS_HEADER_ROW,
            2 + offset,
            severity,
            bg=SEVERITY_CHART_COLORS[severity],
            font=Font(color=SEVERITY_LABEL_FONTS[severity], bold=True),
        )

    for index, (label, status, bg, font_color) in enumerate(CHECKS_ROWS):
        excel_row = CHECKS_FIRST_DATA_ROW + index
        stat_cell(ws, excel_row, 1, label, bg=bg, font=Font(color=font_color, bold=True))
        for offset, severity in enumerate(REPORT_SEVERITIES):
            stat_cell(ws, excel_row, 2 + offset, count_status_severity(status, severity))


def _severity_categories(ws: Worksheet) -> Reference:
    return Reference(ws, min_col=2, max_col=CHECKS_LAST_COL, min_row=CHECKS_HEADER_ROW)


def _failed_by_severity_pie(ws: Worksheet) -> PieChart:
    chart = PieChart()
    chart.add_data(
        Reference(ws, min_col=2, max_col=CHECKS_LAST_COL, min_row=CHECKS_FIRST_DATA_ROW),
        from_rows=True,
        titles_from_data=False,
    )
    chart.set_categories(_severity_categories(ws))
    color_points(chart.series[0], [SEVERITY_CHART_COLORS[s] for s in REPORT_SEVERITIES])
    set_data_labels(chart, show_value=True, position='ctr')
    style_chart(chart, title='Failed checks statistics', width=8, height=7)
    return chart


def _stacked_bar(title: str) -> BarChart:
    chart = BarChart()
    chart.type = 'bar'
    chart.grouping = 'stacked'
    chart.overlap = 100
    set_data_labels(chart, show_value=True, position='ctr')
    style_chart(chart, title=title, width=14, height=5.5)
    return chart


def _severity_vs_status_bar(ws: Worksheet) -> BarChart:
    """One series per status, categories are severities."""
    chart = _stacked_bar('Severity vs Status')
    chart.add_data(
        Reference(
            ws,
            min_col=1,
            max_col=CHECKS_LAST_COL,
            min_row=CHECKS_FIRST_DATA_ROW,
            max_row=CHECKS_LAST_DATA_ROW,
        ),
        from_rows=True,
        titles_from_data=True,
    )
    chart.set_categories(_severity_categories(ws))
    for series, color in zip(chart.series, (CHART_FAIL_COLOR, CHART_PASS_COLOR)):
        color_series(series, color)
    style_value_axes(chart)
    return chart


def _status_vs_severity_bar(ws: Worksheet) -> BarChart:
    """One series per severity, categories are statuses."""
    chart = _stacked_bar('Status vs Severity')
    chart.add_data(
        Reference(
            ws,
            min_col=2,
            max_col=CHECKS_LAST_COL,
            min_row=CHECKS_HEADER_ROW,
            max_row=CHECKS_LAST_DATA_ROW,
        ),
        from_rows=False,
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, min_row=CHECKS_FIRST_DATA_ROW, max_row=CHECKS_LAST_DATA_ROW)
    )
    for series, severity in zip(chart.series, REPORT_SEVERITIES):
        color_series(series, SEVERITY_CHART_COLORS[severity])
    style_value_axes(chart)
    return chart


# ─────────────────────────────────────────────────────────────────────────────
# Workbook assembly
# ─────────────────────────────────────────────────────────────────────────────

def build_workbook(rows: Sequence[ControlRow], standard_name: str, version: str) -> Workbook:
    if not rows:
        raise QuestionnaireError('the schema contains no controls - nothing to write')

    layout = SheetLayout.for_rows(rows)
    wb = Workbook()
    wb.active.title = sheet_title(f'{standard_name} {version}')
    build_questionnaire_sheet(wb.active, rows, layout)
    build_summary_sheet(wb, standard_name, version)
    build_checks_sheet(wb)

    # Structured-reference formulas must be evaluated by Excel on open.
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True
    return wb


def save_workbook(wb: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Orchestration + CLI
# ─────────────────────────────────────────────────────────────────────────────

def generate_questionnaire(
    standard: str,
    compliance_path: Path,
    schemas_dir: Path,
    output_path: Path,
) -> Statistics:
    """Full pipeline: schema -> compliance -> workbook. Returns the statistics."""
    base_name, requested_version = parse_standard_name(standard)
    schema_file = find_schema_file(schemas_dir, base_name)
    version_block, version = load_version_block(schema_file, requested_version)
    log.info(f'Schema     : {schema_file.name}  (version {version})')

    rows = parse_schema(version_block)
    log.info(f'Controls   : {len(rows)} controls loaded from schema')

    compliance = parse_compliance(_load_json(compliance_path), standard)
    if compliance:
        log.info(f'Compliance : {len(compliance)} controls found in report')
    else:
        log.warning(
            f'Warning: standard `{standard}` was not found in the compliance report; '
            f'all controls will be marked {Status.NOT_EVALUATED.value}.',
        )
    apply_compliance(rows, compliance)

    save_workbook(build_workbook(rows, base_name, version), output_path)
    return Statistics.from_rows(rows)


def _load_json(path: Path) -> Mapping[str, Any]:
    try:
        data = json.loads(path.read_text(encoding='utf-8'))
    except OSError as exc:
        raise QuestionnaireError(f'cannot read {path}: {exc}') from exc
    except json.JSONDecodeError as exc:
        raise QuestionnaireError(f'{path} is not valid JSON: {exc}') from exc
    if not isinstance(data, Mapping):
        raise QuestionnaireError(f'{path} must contain a JSON object keyed by region')
    return data


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog='questionnaire',
        description='Generate a standard questionnaire Excel file.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        '--standard',
        required=True,
        metavar='NAME',
        help='Standard name as in the compliance report, e.g. `CIS Controls v7`. '
             'The version suffix selects the version block in the schema file.',
    )
    parser.add_argument(
        '--compliance',
        required=True,
        type=Path,
        metavar='FILE',
        help='Path to the SRE job compliance report JSON.',
    )
    parser.add_argument(
        '--schemas',
        required=True,
        type=Path,
        metavar='DIR',
        help='Repository root containing the standards_schema/ sub-folder.',
    )
    parser.add_argument(
        '--output',
        '-o',
        type=Path,
        default=Path('standard_questionnaire.xlsx'),
        metavar='XLSX',
        help='Output Excel file path (default: %(default)s).',
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format='%(message)s', stream=sys.stdout)
    args = _build_parser().parse_args(argv)

    try:
        if not args.compliance.is_file():
            raise QuestionnaireError(f'compliance file not found: {args.compliance}')
        if not args.schemas.is_dir():
            raise QuestionnaireError(f'schemas directory not found: {args.schemas}')
        stats = generate_questionnaire(args.standard, args.compliance, args.schemas, args.output)
    except QuestionnaireError as exc:
        log.error(f'Error: {exc}')
        return 1

    log.info(stats.render(args.output))
    return 0


if __name__ == '__main__':
    sys.exit(main())
