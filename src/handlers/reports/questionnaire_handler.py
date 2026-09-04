"""
Standard Questionnaire report.

Builds an Excel workbook that maps every control of a security standard to its
audit status, coverage and the rules that evaluate it.

Sheets
------
  1. 'Introduction'         - report explanation, scope and workbook guide.
  2. 'Compliance'           - one row per control (Excel table + conditional
     formatting, so manual edits stay color-coded).
  3. 'Summary'              - formula-driven statistics table + two pie charts
     and two stacked charts.
  ('_ChartData' is a hidden helper sheet feeding the bar-of-pie chart.)

Scope
-----
  jobs        - results of one single scan job.
  accumulated - the latest accumulated state of a tenant.

File names
----------
  jobs-<standard>-<job id>-questionnaire.xlsx
  accumulated-<standard>-<tenant name>-questionnaire.xlsx

  <standard> is the standard name with its version slugified with
  underscores, <job id> is the last segment of the job UUID.

Data sources
------------
  Customer metadata (LM) : the full list of the standard controls, per control
    severity and the rules that are mapped to each control.
  Scan results           : which of those rules succeeded / failed / were not
    evaluated.

Currently metadata does not provide the standard hierarchy (sections,
sub-sections, control titles) therefore those columns are omitted. As soon as
such data appears it is enough to populate the corresponding ``ControlRow``
attributes - the sheet layout picks them up automatically.

Multi-region aggregation
------------------------
  failed        = union over all regions (worst case wins)
  successful    = union minus failed
  not_evaluated = union minus failed and successful

Status logic
------------
  Fail          - at least one rule failed
  Pass          - every rule succeeded
  Not Evaluated - no failure but at least one rule unevaluated (or no rules)
"""

import io
import re
from collections.abc import Iterable, Mapping as MappingABC, Sequence
from dataclasses import dataclass, field
from enum import Enum
from http import HTTPStatus
from itertools import chain
from typing import Any

from modular_sdk.services.tenant_service import TenantService
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, ProjectedPieChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from typing_extensions import Self

from handlers import AbstractHandler, Mapping
from helpers import deep_get
from helpers.constants import (
    DEPRECATED_RULE_SUFFIX,
    Cloud,
    Endpoint,
    HTTPMethod,
    ReportFormat,
    Severity,
)
from helpers.lambda_response import build_response
from helpers.log_helper import get_logger
from helpers.reports import Standard
from services import SP, modular_helpers
from services.coverage_service import calculate_controls_coverages
from services.job_service import JobService
from services.license_service import LicenseService
from services.metadata import Metadata
from services.platform_service import PlatformService
from services.report_service import ReportResponse, ReportService
from services.sharding import ShardsCollection
from validators.swagger_request_models import (
    JobQuestionnaireReportGetModel,
    TenantQuestionnaireReportGetModel,
)
from validators.utils import validate_kwargs


_LOG = get_logger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEX_HEADER_BG = '4472C4'
HEX_HEADER_FONT = 'FFFFFF'
HEX_GREEN_PASTEL = 'C6EFCE'
HEX_GREEN_FONT = '006100'
HEX_RED_PASTEL = 'FFC7CE'
HEX_RED_FONT = '9C0006'
HEX_GRAY = 'BFBFBF'
HEX_GRAY_FONT = '404040'
HEX_BORDER_BLUE = '9DC3E6'

CHART_FAIL_COLOR = 'ED7D31'
CHART_PASS_COLOR = '70AD47'
CHART_NA_COLOR = 'A5A5A5'

#: Severities that the questionnaire actually uses in charts/tables.
REPORT_SEVERITIES: tuple[str, ...] = (
    Severity.HIGH.value,
    Severity.MEDIUM.value,
    Severity.LOW.value,
)
SEVERITY_CHART_COLORS = {'High': 'FF3300', 'Medium': 'FFC000', 'Low': 'BDD7EE'}
SEVERITY_LABEL_FONTS = {'High': 'FFFFFF', 'Medium': '000000', 'Low': '000000'}
#: Conditional-format fills for the Severity column of the questionnaire.
SEVERITY_CELL_FILLS = {
    'High': 'FF3300',
    'Medium': 'FFCC00',
    'Low': 'FFFF99',
    'Info': 'A6C9EC',
}

BORDER_SIDE = Side(style='thin', color=HEX_BORDER_BLUE)
BLUE_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)
_THIN = Side(style='thin')
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WHITE_FILL = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')

ALIGN_LEFT = Alignment(horizontal='left', vertical='top')
ALIGN_CENTER = Alignment(horizontal='center', vertical='top')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)

PERCENT_FORMAT = '0%'
HEADER_ROW = 1
FIRST_DATA_ROW = 2

TABLE_NAME = 'StandardQuestionnaire'
COMPLIANCE_SHEET = 'Compliance'
CHART_DATA_SHEET = '_ChartData'
CHART_FONT = 'Aptos Narrow'
CHART_TITLE_SIZE = 1600
CHART_LEGEND_SIZE = 900

_NATURAL_CHUNKS = re.compile(r'(\d+)')

#: Header names used by :class:`SheetLayout`
H_SECTION = 'Section'
H_TITLE = 'Title'
H_CONTROL = 'Control'
H_SEVERITY = 'Severity'
H_STATUS = 'Status'
H_CONTROL_TITLE = 'Control Title'
H_COVERAGE = 'Coverage'
H_RULE_IDS = 'Rule IDs'


# ---------------------------------------------------------------------------
# Domain model
# ---------------------------------------------------------------------------


class Status(str, Enum):
    PASS = 'Pass'
    FAIL = 'Fail'
    NOT_EVALUATED = 'Not Evaluated'


def slugify(value: str) -> str:
    """'CIS Kubernetes Benchmark v1.7.0' -> 'cis_kubernetes_benchmark_v1_7_0'."""
    return re.sub(r'[^a-z0-9]+', '_', value.lower()).strip('_')


def short_job_id(job_id: str) -> str:
    """Last segment of the job UUID, i.e. '...-9f2c1d4e' -> '9f2c1d4e'."""
    return job_id.rsplit('-', maxsplit=1)[-1]


def is_deprecated_rule(rule_name: str) -> bool:
    # TODO: update when rule deprecation logic changes
    return rule_name.endswith(DEPRECATED_RULE_SUFFIX)


class ReportScope(str, Enum):
    """Which questionnaire flavour is being built."""

    JOB = 'jobs'
    ACCUMULATED = 'accumulated'

    @property
    def label(self) -> str:
        match self:
            case ReportScope.JOB:
                return 'Job questionnaire'
            case _:
                return 'Accumulated (tenant) questionnaire'

    @property
    def description(self) -> str:
        match self:
            case ReportScope.JOB:
                return (
                    'This workbook is a job questionnaire. It reflects the '
                    'results of one single scan job of one tenant or '
                    'Kubernetes platform. Controls that the job did not '
                    'evaluate are reported as Not Evaluated even if another '
                    'job has already covered them.'
                )
            case _:
                return (
                    'This workbook is an accumulated (tenant) questionnaire. '
                    'It reflects the latest accumulated state of the tenant, '
                    'meaning the most recent result of every rule collected '
                    'over all of the scans of that tenant rather than a '
                    'single scan job.'
                )

    @property
    def filename_note(self) -> str:
        match self:
            case ReportScope.JOB:
                return (
                    'The file is named '
                    'jobs-<standard>-<job id>-questionnaire.xlsx, where '
                    '<standard> is the standard name with its version and '
                    '<job id> is the last segment of the job UUID.'
                )
            case _:
                return (
                    'The file is named '
                    'accumulated-<standard>-<tenant name>-questionnaire.xlsx, '
                    'where <standard> is the standard name with its version.'
                )

    def filename(self, standard: Standard, entity: str) -> str:
        return (
            f'{self.value}-{slugify(standard.full_name)}-{entity}'
            f'-questionnaire.xlsx'
        )


STATUS_FILL_COLORS = {
    Status.PASS.value: HEX_GREEN_PASTEL,
    Status.FAIL.value: HEX_RED_PASTEL,
    Status.NOT_EVALUATED.value: HEX_GRAY,
}
STATUS_FONT_COLORS = {
    Status.PASS.value: HEX_GREEN_FONT,
    Status.FAIL.value: HEX_RED_FONT,
    Status.NOT_EVALUATED.value: HEX_GRAY_FONT,
}


def short_rule_id(rule_id: str) -> str:
    """'ecc-aws-001-some_name' -> '001'; anything else is returned unchanged."""
    parts = rule_id.split('-')
    return parts[2] if len(parts) >= 3 and parts[2].isdigit() else rule_id


def natural_key(value: str) -> tuple:
    """'1.10' sorts after '1.2'."""
    return tuple(
        int(chunk) if chunk.isdigit() else chunk
        for chunk in _NATURAL_CHUNKS.split(value)
    )


@dataclass
class ControlRow:
    """One control of the standard, enriched with compliance results."""

    control_id: str
    #: Optional hierarchy data. Not provided by metadata yet, but the sheet
    #: layout supports it as soon as it appears.
    section_id: str = ''
    section_title: str = ''
    #: (id, title) pairs between the top-level section and the control.
    sub_sections: list[tuple[str, str]] = field(default_factory=list)
    control_title: str = ''
    coverage: float = 0.0
    severity: str = ''
    successful_rules: list[str] = field(default_factory=list)
    failed_rules: list[str] = field(default_factory=list)
    not_evaluated_rules: list[str] = field(default_factory=list)

    @property
    def status(self) -> Status:
        if self.failed_rules:
            return Status.FAIL
        if self.not_evaluated_rules or not self.successful_rules:
            return Status.NOT_EVALUATED
        return Status.PASS

    @property
    def rule_short_ids(self) -> str:
        ids = {
            short_rule_id(rule)
            for rule in chain(
                self.successful_rules,
                self.failed_rules,
                self.not_evaluated_rules,
            )
        }
        return ', '.join(sorted(ids))


@dataclass
class ControlCompliance:
    """Compliance results for one control, aggregated over all regions."""

    coverage: float | None = None
    severity: str | None = None
    successful: set[str] = field(default_factory=set)
    failed: set[str] = field(default_factory=set)
    not_evaluated: set[str] = field(default_factory=set)

    def absorb(
        self,
        *,
        coverage: float | None = None,
        severity: str | None = None,
        successful: Iterable[str] = (),
        failed: Iterable[str] = (),
        not_evaluated: Iterable[str] = (),
    ) -> None:
        """Merge one region's results for this control (first values win)."""
        if self.coverage is None and coverage is not None:
            self.coverage = float(coverage)
        if self.severity is None and severity is not None:
            self.severity = str(getattr(severity, 'value', severity))
        self.successful.update(successful)
        self.failed.update(failed)
        self.not_evaluated.update(not_evaluated)

    def resolve_overlaps(self) -> None:
        """A rule that failed anywhere counts as failed; then as successful."""
        self.successful -= self.failed
        self.not_evaluated -= self.failed | self.successful


# ---------------------------------------------------------------------------
# Excel helpers - styles, formulas, charts
# ---------------------------------------------------------------------------


def fill(color: str) -> PatternFill:
    return PatternFill('solid', start_color=color, end_color=color)


def table_ref(column: str) -> str:
    return f'{TABLE_NAME}[{column}]'


STATUS_REF = table_ref(H_STATUS)
SEVERITY_REF = table_ref(H_SEVERITY)
COVERAGE_REF = table_ref(H_COVERAGE)


def count_status(status: Status) -> str:
    return f'=COUNTIF({STATUS_REF},"{status.value}")'


def count_status_severity(status: Status, severity: str) -> str:
    return (
        f'=COUNTIFS({STATUS_REF},"{status.value}",{SEVERITY_REF},"{severity}")'
    )


def stat_cell(
    ws: Worksheet,
    row: int,
    column: int,
    value: Any = None,
    *,
    bg: str | None = None,
    font: Font | None = None,
    align: Alignment = ALIGN_CENTER,
):
    """Bordered cell used by the Summary statistics table."""
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = THIN_BORDER
    cell.alignment = align
    if bg:
        cell.fill = fill(bg)
    if font:
        cell.font = font
    return cell


def write_sheet_title(ws: Worksheet, text: str) -> None:
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, size=14)


def set_column_widths(ws: Worksheet, widths: MappingABC[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def autofit_columns(
    ws: Worksheet, *, max_row: int, limit: int = 50, minimum: int = 10
) -> None:
    for column in ws.iter_cols(min_row=1, max_row=max_row):
        longest = max(
            (len(str(c.value)) for c in column if c.value is not None),
            default=0,
        )
        letter = get_column_letter(column[0].column)
        ws.column_dimensions[letter].width = max(
            minimum, min(longest + 2, limit)
        )


def add_text_conditional_format(
    ws: Worksheet,
    column: int,
    last_row: int,
    fills: MappingABC[str, str],
    fonts: MappingABC[str, str] | None = None,
) -> None:
    """Color cells by their own text, so manual edits stay color-coded."""
    if last_row < FIRST_DATA_ROW:
        return
    letter = get_column_letter(column)
    cell_range = f'{letter}{FIRST_DATA_ROW}:{letter}{last_row}'
    for text, bg in fills.items():
        font_color = (fonts or {}).get(text)
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'UPPER(${letter}{FIRST_DATA_ROW})="{text.upper()}"'],
                fill=fill(bg),
                font=Font(color=font_color) if font_color else None,
            ),
        )


def _chart_text(size: int) -> RichText:
    return RichText(
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(
                        sz=size, latin=DrawingFont(typeface=CHART_FONT)
                    )
                )
            )
        ]
    )


def style_chart(chart, *, title: str, width: float = 10, height: float = 6):
    chart.title = title
    chart.graphical_properties = GraphicalProperties(
        ln=LineProperties(noFill=True)
    )
    chart.roundedCorners = False
    if chart.title is not None:
        chart.title.overlay = False
        chart.title.txPr = _chart_text(CHART_TITLE_SIZE)
    chart.legend.position = 'r'
    chart.legend.overlay = False
    chart.legend.txPr = _chart_text(CHART_LEGEND_SIZE)
    chart.width = width
    chart.height = height


def set_data_labels(
    chart,
    *,
    show_value: bool = False,
    show_percent: bool = False,
    position: str | None = None,
) -> None:
    labels = DataLabelList()
    labels.showVal = show_value
    labels.showPercent = show_percent
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    if position:
        labels.dLblPos = position
    chart.dataLabels = labels


def color_points(series, colors: Sequence[str]) -> None:
    """Color individual pie slices, in data order."""
    for idx, color in enumerate(colors):
        point = DataPoint(idx=idx)
        point.graphicalProperties = GraphicalProperties(solidFill=color)
        series.data_points.append(point)


def color_series(series, color: str) -> None:
    series.graphicalProperties = GraphicalProperties(solidFill=color)


def style_value_axes(chart) -> None:
    for axis in (chart.x_axis, chart.y_axis):
        axis.delete = False
        axis.txPr = _chart_text(CHART_LEGEND_SIZE)
    chart.y_axis.tickLblPos = 'low'


# ---------------------------------------------------------------------------
# Introduction sheet
# ---------------------------------------------------------------------------

INTRODUCTION_SHEET = 'Introduction'
INTRODUCTION_LAST_COL = 6
INTRODUCTION_HEADER_BG = 'FF00549F'
INTRODUCTION_NOTICE_BG = 'FF7F7F7F'
INTRODUCTION_HEADER_FONT = 'Gill Sans MT'
INTRODUCTION_BODY_FONT = 'Aptos Narrow'

INTRO_TITLE_HEIGHT = 30.0
INTRO_SECTION_HEIGHT = 19.9
INTRO_LEAD_HEIGHT = 14.45
INTRO_BODY_HEIGHT = 34.15
INTRO_BULLET_HEIGHT = 16.9
INTRO_NOTICE_HEIGHT = 31.0


class _IntroSheetWriter:
    """Sequential writer: sections can be added without fixed row numbers."""

    def __init__(self, ws: Worksheet) -> None:
        self._ws = ws
        self._row = 1

    def _write(
        self,
        text: str,
        *,
        span: int,
        height: float,
        font: Font,
        alignment: Alignment,
        background: str | None = None,
    ) -> None:
        first, last = self._row, self._row + span - 1
        self._ws.merge_cells(
            start_row=first,
            start_column=1,
            end_row=last,
            end_column=INTRODUCTION_LAST_COL,
        )
        cell = self._ws.cell(row=first, column=1, value=text)
        cell.font = font
        cell.alignment = alignment
        if background:
            cell.fill = fill(background)
        for row in range(first, last + 1):
            self._ws.row_dimensions[row].height = height
        self._row = last + 1

    def section(
        self,
        text: str,
        *,
        background: str = INTRODUCTION_HEADER_BG,
        height: float = INTRO_SECTION_HEIGHT,
    ) -> None:
        self._write(
            text,
            span=1,
            height=height,
            font=Font(
                name=INTRODUCTION_HEADER_FONT,
                size=13,
                bold=True,
                color=HEX_HEADER_FONT,
            ),
            alignment=ALIGN_HEADER,
            background=background,
        )

    def title(self, text: str) -> None:
        self.section(text, height=INTRO_TITLE_HEIGHT)

    def body(
        self, text: str, *, span: int = 2, height: float = INTRO_BODY_HEIGHT
    ) -> None:
        self._write(
            text,
            span=span,
            height=height,
            font=Font(name=INTRODUCTION_BODY_FONT, size=11),
            alignment=Alignment(
                horizontal='left', vertical='center', wrap_text=True
            ),
        )

    def bullet(self, text: str) -> None:
        self.body(f'\u2022  {text}', span=1, height=INTRO_BULLET_HEIGHT)

    def blank(self) -> None:
        self._row += 1


def build_introduction_sheet(
    wb: Workbook, standard: Standard, scope: ReportScope
) -> None:
    """Create the report guide shown before the Compliance and Summary sheets."""
    ws = wb.create_sheet(INTRODUCTION_SHEET, 0)
    writer = _IntroSheetWriter(ws)

    writer.title('Introduction')
    writer.body(
        f'This workbook is a detailed compliance questionnaire for '
        f'{standard.full_name}. It maps every control of the selected '
        'security standard to the coverage, severity, evaluation status, and '
        'rules reported by SRE to support control coverage and rule-result '
        'review.',
        span=3,
        height=INTRO_LEAD_HEIGHT,
    )
    writer.blank()

    writer.section('I. Scope')
    writer.body(
        f'Report type: {scope.label}.', span=1, height=INTRO_BULLET_HEIGHT
    )
    writer.body(scope.description)
    writer.body(scope.filename_note)
    writer.blank()

    writer.section('II. Structure', height=25.15)
    writer.body(
        'The spreadsheet includes three visible sheets and one hidden helper '
        'sheet:',
        span=1,
        height=24.6,
    )
    writer.bullet(
        f'{COMPLIANCE_SHEET} - one row per control, including its severity, '
        'status, coverage, and associated rule IDs.'
    )
    writer.bullet(
        'Summary - formula-driven status and severity statistics, coverage '
        'metrics, and charts.'
    )
    writer.bullet(
        '_ChartData - hidden helper cells used by the Applicable Coverage '
        'chart.'
    )
    writer.blank()

    writer.section('III. Description')
    writer.body(
        f'{COMPLIANCE_SHEET} contains the full list of the standard controls '
        'taken from the customer metadata. Coverage and severity are shown '
        'alongside the rule IDs that passed, failed, or were not evaluated.'
    )
    writer.body(
        'Summary contains the Total Coverage and Applicable Coverage pies, '
        'plus Severity vs Status and Status vs Severity stacked-bar charts. '
        'Its statistics table is formula-driven and updates from the '
        'StandardQuestionnaire table when Excel recalculates the workbook.'
    )
    writer.body(
        'Status is derived from rule results, not from the coverage number: '
        'Fail means at least one rule failed; Pass means every rule succeeded; '
        'Not Evaluated means there was no failure but at least one rule was '
        'unevaluated or no rule succeeded.'
    )
    writer.body(
        'SRE Applicable Coverage excludes Not Evaluated controls, while Total '
        'Coverage averages all controls.'
    )
    writer.blank()

    writer.section('Notice', background=INTRODUCTION_NOTICE_BG)
    writer.body(
        'When a scan contains multiple regions, rule lists are unioned across '
        'regions and failed rules take precedence over successful or '
        'unevaluated results. Coverage is calculated once for the whole '
        'account. Do not delete the hidden _ChartData sheet while charts are '
        'in use.',
        span=3,
        height=INTRO_NOTICE_HEIGHT,
    )

    set_column_widths(ws, {'A': 22.3, 'B': 35.6, 'C': 15.6, 'D': 80.0})
    ws.sheet_properties.tabColor = INTRODUCTION_HEADER_BG


# ---------------------------------------------------------------------------
# Questionnaire sheet
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SheetLayout:
    """Column geometry of the questionnaire sheet (1-based indices)."""

    sub_depth: int = 0
    with_sections: bool = False
    with_control_title: bool = False

    @classmethod
    def for_rows(cls, rows: Sequence[ControlRow]) -> 'SheetLayout':
        return cls(
            sub_depth=max((len(r.sub_sections) for r in rows), default=0),
            with_sections=any(r.section_id or r.section_title for r in rows),
            with_control_title=any(r.control_title for r in rows),
        )

    @property
    def sub_headers(self) -> list[str]:
        if self.sub_depth == 0:
            return []
        if self.sub_depth == 1:
            return ['Sub-section', 'Sub-title']
        return [
            header
            for level in range(2, self.sub_depth + 2)
            for header in (
                f'Sub-section (l{level})',
                f'Sub-title (l{level})',
            )
        ]

    @property
    def headers(self) -> list[str]:
        headers = []
        if self.with_sections:
            headers += [H_SECTION, H_TITLE]
        headers += self.sub_headers
        headers += [H_CONTROL, H_SEVERITY, H_STATUS]
        if self.with_control_title:
            headers.append(H_CONTROL_TITLE)
        headers += [H_COVERAGE, H_RULE_IDS]
        return headers

    @property
    def width(self) -> int:
        return len(self.headers)

    def index(self, header: str) -> int:
        return self.headers.index(header) + 1

    @property
    def severity(self) -> int:
        return self.index(H_SEVERITY)

    @property
    def status(self) -> int:
        return self.index(H_STATUS)

    @property
    def control_title(self) -> int:
        return self.index(H_CONTROL_TITLE) if self.with_control_title else -1

    @property
    def coverage(self) -> int:
        return self.index(H_COVERAGE)

    def alignment(self, column: int) -> Alignment:
        if column == self.status:
            return ALIGN_CENTER
        if column == self.control_title:
            return ALIGN_WRAP
        return ALIGN_LEFT

    def row_values(self, row: ControlRow) -> list[Any]:
        """Cell values for one control, always ``self.width`` long."""
        values: list[Any] = []
        if self.with_sections:
            values += [row.section_id or None, row.section_title or None]
        for level in range(self.sub_depth):
            pair = (
                row.sub_sections[level]
                if level < len(row.sub_sections)
                else (None, None)
            )
            values.extend(pair)
        values += [
            row.control_id,
            row.severity or None,
            row.status.value,
        ]
        if self.with_control_title:
            values.append(row.control_title or None)
        values += [row.coverage, row.rule_short_ids or None]
        return values


def build_questionnaire_sheet(
    ws: Worksheet, rows: Sequence[ControlRow], layout: SheetLayout
) -> None:
    for column, header in enumerate(layout.headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=column, value=header)
        cell.fill = fill(HEX_HEADER_BG)
        cell.font = Font(color=HEX_HEADER_FONT, bold=True)
        cell.alignment = ALIGN_HEADER
        cell.border = BLUE_BORDER
    ws.row_dimensions[HEADER_ROW].height = 30

    for offset, row in enumerate(rows):
        excel_row = FIRST_DATA_ROW + offset
        for column, value in enumerate(layout.row_values(row), start=1):
            cell = ws.cell(row=excel_row, column=column, value=value)
            cell.alignment = layout.alignment(column)
            cell.border = BLUE_BORDER
            cell.fill = WHITE_FILL
            if column == layout.coverage:
                cell.number_format = PERCENT_FORMAT

    last_row = FIRST_DATA_ROW + len(rows) - 1
    _add_questionnaire_table(ws, layout, last_row)
    add_text_conditional_format(
        ws, layout.severity, last_row, SEVERITY_CELL_FILLS
    )
    add_text_conditional_format(
        ws, layout.status, last_row, STATUS_FILL_COLORS, STATUS_FONT_COLORS
    )
    _add_coverage_totals(ws, layout, last_row)

    autofit_columns(ws, max_row=last_row)
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1).coordinate


def _add_questionnaire_table(
    ws: Worksheet, layout: SheetLayout, last_row: int
) -> None:
    ref = f'A{HEADER_ROW}:{get_column_letter(layout.width)}{last_row}'
    table = Table(displayName=TABLE_NAME, ref=ref)
    table.autoFilter = AutoFilter(ref=ref)
    table.tableColumns = [
        TableColumn(id=index, name=header)
        for index, header in enumerate(layout.headers, start=1)
    ]
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_coverage_totals(
    ws: Worksheet, layout: SheetLayout, last_row: int
) -> None:
    """Coverage averages written *outside* the table so they are not filtered."""
    totals = (
        (
            'SRE Applicable Coverage',
            f'=IFERROR(AVERAGEIF({STATUS_REF},'
            f'"<>{Status.NOT_EVALUATED.value}",{COVERAGE_REF}),0)',
        ),
        ('Total Coverage', f'=IFERROR(AVERAGE({COVERAGE_REF}),0)'),
    )
    for offset, (label, formula) in enumerate(totals, start=2):
        excel_row = last_row + offset
        label_cell = ws.cell(row=excel_row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = ALIGN_LEFT

        value_cell = ws.cell(
            row=excel_row, column=layout.coverage, value=formula
        )
        value_cell.font = Font(bold=True)
        value_cell.alignment = ALIGN_LEFT
        value_cell.number_format = PERCENT_FORMAT


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

SUMMARY_TABLE_ROW = 3
#: (label, status, fill, font color, split by severity)
SUMMARY_ROWS: tuple[tuple[str, Status, str, str, bool], ...] = (
    ('Fail', Status.FAIL, HEX_RED_PASTEL, HEX_RED_FONT, True),
    ('Pass', Status.PASS, HEX_GREEN_PASTEL, HEX_GREEN_FONT, True),
    ('Not Applicable', Status.NOT_EVALUATED, HEX_GRAY, HEX_GRAY_FONT, False),
)
SUMMARY_FIRST_DATA_ROW = SUMMARY_TABLE_ROW + 1
SUMMARY_LAST_STATUS_ROW = SUMMARY_FIRST_DATA_ROW + 1
SUMMARY_FIRST_SEVERITY_COL = 3
SUMMARY_LAST_SEVERITY_COL = (
    SUMMARY_FIRST_SEVERITY_COL + len(REPORT_SEVERITIES) - 1
)
SUMMARY_COVERAGE_CHART_ROW = SUMMARY_TABLE_ROW + len(SUMMARY_ROWS) + 3


def build_summary_sheet(
    wb: Workbook, standard_name: str, version: str
) -> None:
    """Formula-driven statistics table plus four charts."""
    ws = wb.create_sheet('Summary')
    helper = wb.create_sheet(CHART_DATA_SHEET)
    helper.sheet_state = 'hidden'

    write_sheet_title(ws, f'Summary by {standard_name} {version}')
    _write_summary_table(ws)
    _write_bar_of_pie_data(helper)

    ws.add_chart(_total_coverage_pie(ws), f'A{SUMMARY_COVERAGE_CHART_ROW}')
    ws.add_chart(
        _applicable_coverage_pie(helper), f'F{SUMMARY_COVERAGE_CHART_ROW}'
    )
    ws.add_chart(_severity_vs_status_bar(ws), 'A22')
    ws.add_chart(_status_vs_severity_bar(ws), 'A34')

    set_column_widths(
        ws, {'A': 18, 'B': 12, 'C': 10, 'D': 10, 'E': 10, 'F': 3}
    )


def _write_summary_table(ws: Worksheet) -> None:
    stat_cell(ws, SUMMARY_TABLE_ROW, 1, font=Font(bold=True))
    stat_cell(ws, SUMMARY_TABLE_ROW, 2, 'Controls', font=Font(bold=True))
    for offset, severity in enumerate(REPORT_SEVERITIES):
        stat_cell(
            ws,
            SUMMARY_TABLE_ROW,
            3 + offset,
            severity,
            bg=SEVERITY_CHART_COLORS[severity],
            font=Font(color=SEVERITY_LABEL_FONTS[severity], bold=True),
        )

    for index, (label, status, bg, font_color, by_severity) in enumerate(
        SUMMARY_ROWS, start=1
    ):
        excel_row = SUMMARY_TABLE_ROW + index
        stat_cell(
            ws, excel_row, 1, label, bg=bg, font=Font(color=font_color, bold=True)
        )
        stat_cell(ws, excel_row, 2, count_status(status))
        for offset, severity in enumerate(REPORT_SEVERITIES):
            value = (
                count_status_severity(status, severity) if by_severity else '-'
            )
            stat_cell(ws, excel_row, 3 + offset, value)

    total_row = SUMMARY_TABLE_ROW + len(SUMMARY_ROWS) + 1
    stat_cell(ws, total_row, 1, 'Total', font=Font(bold=True))
    stat_cell(ws, total_row, 2, f'=ROWS({STATUS_REF})', font=Font(bold=True))
    for offset in range(len(REPORT_SEVERITIES)):
        stat_cell(ws, total_row, 3 + offset)


def _write_bar_of_pie_data(helper: Worksheet) -> None:
    entries = [('Pass', count_status(Status.PASS))] + [
        (severity, count_status_severity(Status.FAIL, severity))
        for severity in REPORT_SEVERITIES
    ]
    for index, (label, formula) in enumerate(entries, start=1):
        helper.cell(row=index, column=1, value=label)
        helper.cell(row=index, column=2, value=formula)


def _total_coverage_pie(ws: Worksheet) -> PieChart:
    first, last = (
        SUMMARY_FIRST_DATA_ROW,
        SUMMARY_TABLE_ROW + len(SUMMARY_ROWS),
    )
    chart = PieChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=first, max_row=last),
        titles_from_data=False,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    color_points(
        chart.series[0],
        (CHART_FAIL_COLOR, CHART_PASS_COLOR, CHART_NA_COLOR),
    )
    set_data_labels(chart, show_percent=True)
    style_chart(chart, title='Total Coverage')
    return chart


def _applicable_coverage_pie(helper: Worksheet) -> ProjectedPieChart:
    last = 1 + len(REPORT_SEVERITIES)
    chart = ProjectedPieChart()
    chart.type = 'bar'
    chart.splitType = 'pos'
    chart.splitPos = len(REPORT_SEVERITIES)
    chart.add_data(
        Reference(helper, min_col=2, min_row=1, max_row=last),
        titles_from_data=False,
    )
    chart.set_categories(
        Reference(helper, min_col=1, min_row=1, max_row=last)
    )

    series = chart.series[0]
    color_series(series, CHART_FAIL_COLOR)  # combined 'other' slice
    color_points(
        series,
        (CHART_PASS_COLOR, *(SEVERITY_CHART_COLORS[s] for s in REPORT_SEVERITIES)),
    )
    set_data_labels(chart, show_percent=True)
    style_chart(chart, title='Applicable Coverage')
    return chart


def _severity_categories(ws: Worksheet) -> Reference:
    return Reference(
        ws,
        min_col=SUMMARY_FIRST_SEVERITY_COL,
        max_col=SUMMARY_LAST_SEVERITY_COL,
        min_row=SUMMARY_TABLE_ROW,
    )


def _stacked_bar(title: str) -> BarChart:
    chart = BarChart()
    chart.type = 'bar'
    chart.grouping = 'stacked'
    chart.overlap = 100
    set_data_labels(chart, show_value=True, position='ctr')
    style_chart(chart, title=title, width=14, height=5.5)
    return chart


def _severity_vs_status_bar(ws: Worksheet) -> BarChart:
    """One series per status, categories are severities."""
    chart = _stacked_bar('Severity vs Status')
    chart.add_data(
        Reference(
            ws,
            min_col=SUMMARY_FIRST_SEVERITY_COL,
            max_col=SUMMARY_LAST_SEVERITY_COL,
            min_row=SUMMARY_FIRST_DATA_ROW,
            max_row=SUMMARY_LAST_STATUS_ROW,
        ),
        from_rows=True,
        titles_from_data=False,
    )
    chart.set_categories(_severity_categories(ws))
    for row, series, color in zip(
        range(SUMMARY_FIRST_DATA_ROW, SUMMARY_LAST_STATUS_ROW + 1),
        chart.series,
        (CHART_FAIL_COLOR, CHART_PASS_COLOR),
    ):
        series.title = SeriesLabel(strRef=StrRef(f=f"'{ws.title}'!$A${row}"))
        color_series(series, color)
    style_value_axes(chart)
    return chart


def _status_vs_severity_bar(ws: Worksheet) -> BarChart:
    """One series per severity, categories are statuses."""
    chart = _stacked_bar('Status vs Severity')
    chart.add_data(
        Reference(
            ws,
            min_col=SUMMARY_FIRST_SEVERITY_COL,
            max_col=SUMMARY_LAST_SEVERITY_COL,
            min_row=SUMMARY_TABLE_ROW,
            max_row=SUMMARY_LAST_STATUS_ROW,
        ),
        from_rows=False,
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=1,
            min_row=SUMMARY_FIRST_DATA_ROW,
            max_row=SUMMARY_LAST_STATUS_ROW,
        )
    )
    for series, severity in zip(chart.series, REPORT_SEVERITIES):
        color_series(series, SEVERITY_CHART_COLORS[severity])
    style_value_axes(chart)
    return chart


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------


def build_workbook(
    rows: Sequence[ControlRow], standard: Standard, scope: ReportScope
) -> Workbook:
    layout = SheetLayout.for_rows(rows)
    wb = Workbook()
    compliance_ws = wb.active
    compliance_ws.title = COMPLIANCE_SHEET
    build_questionnaire_sheet(compliance_ws, rows, layout)
    build_introduction_sheet(wb, standard, scope)
    build_summary_sheet(wb, standard.name, standard.version or '')

    # Structured-reference formulas must be evaluated by Excel on open.
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True
    return wb


# ---------------------------------------------------------------------------
# Handler
# ---------------------------------------------------------------------------


class QuestionnaireHandler(AbstractHandler):
    def __init__(
        self,
        tenant_service: TenantService,
        job_service: JobService,
        report_service: ReportService,
        license_service: LicenseService,
        platform_service: PlatformService,
    ):
        self._tenant_service = tenant_service
        self._job_service = job_service
        self._report_service = report_service
        self._license_service = license_service
        self._platform_service = platform_service

    @classmethod
    def build(cls) -> Self:
        return cls(
            tenant_service=SP.modular_client.tenant_service(),
            job_service=SP.job_service,
            report_service=SP.report_service,
            license_service=SP.license_service,
            platform_service=SP.platform_service,
        )

    @property
    def mapping(self) -> Mapping:
        return {
            Endpoint.REPORTS_QUESTIONNAIRE_JOBS_JOB_ID: {
                HTTPMethod.GET: self.get_by_job
            },
            Endpoint.REPORTS_QUESTIONNAIRE_TENANTS_TENANT_NAME: {
                HTTPMethod.GET: self.get_by_tenant
            },
        }

    @validate_kwargs
    def get_by_job(self, event: JobQuestionnaireReportGetModel, job_id: str):
        job = self._job_service.get_nullable(job_id)
        if not job:
            return build_response(
                content='The request job not found',
                code=HTTPStatus.NOT_FOUND,
            )

        if job.is_platform_job:
            platform = self._platform_service.get_nullable(job.platform_id)
            if not platform or (
                event.customer and platform.customer != event.customer
            ):
                return build_response(
                    content='Job platform not found',
                    code=HTTPStatus.NOT_FOUND,
                )
            collection = self._report_service.platform_job_collection(
                platform, job
            )
            collection.meta = self._report_service.fetch_meta(platform)
            cloud = Cloud.KUBERNETES
        else:
            tenant = self._tenant_service.get(job.tenant_name)
            tenant = modular_helpers.assert_tenant_valid(tenant, event.customer)
            collection = self._report_service.job_collection(tenant, job)
            collection.meta = self._report_service.fetch_meta(tenant)
            cloud = modular_helpers.tenant_cloud(tenant)

        metadata = self._license_service.get_customer_metadata(
            event.customer_id
        )
        for standard in metadata.domains['AWS'].full_cov.keys():
            print(standard.full_name)
        print()
        for standard in metadata.domains['KUBERNETES'].full_cov.keys():
            print(standard.full_name)
        full_cov = metadata.domain(cloud).full_cov
        standard = self.resolve_standard(full_cov, event.standard)
        if standard is None:
            return build_response(
                content=f'Standard {event.standard} not found in metadata',
                code=HTTPStatus.NOT_FOUND,
            )

        buffer = self._build_questionnaire(
            collection=collection,
            metadata=metadata,
            cloud=cloud,
            standard=standard,
            full_controls=full_cov[standard],
            scope=ReportScope.JOB,
        )

        url = self._report_service.one_time_url(
            buffer, ReportScope.JOB.filename(standard, short_job_id(job.id))
        )
        response = ReportResponse(job, url, fmt=ReportFormat.XLSX)
        return build_response(content=response.dict())

    @validate_kwargs
    def get_by_tenant(
        self, event: TenantQuestionnaireReportGetModel, tenant_name: str
    ):
        tenant = self._tenant_service.get(tenant_name)
        if not tenant:
            return build_response(
                content=f"Tenant {tenant_name!r} not found", code=HTTPStatus.NOT_FOUND
            )
        modular_helpers.assert_tenant_valid(tenant, event.customer)
        cloud = modular_helpers.tenant_cloud(tenant)
        collection = self._report_service.tenant_latest_collection(tenant)
        collection.meta = self._report_service.fetch_meta(tenant)
        metadata = self._license_service.get_customer_metadata(
            event.customer_id
        )
        full_cov = metadata.domain(cloud).full_cov
        standard = self.resolve_standard(full_cov, event.standard)
        if standard is None:
            return build_response(
                content=f'Standard {event.standard} not found in metadata',
                code=HTTPStatus.NOT_FOUND,
            )

        buffer = self._build_questionnaire(
            collection=collection,
            metadata=metadata,
            cloud=cloud,
            standard=standard,
            full_controls=full_cov[standard],
            scope=ReportScope.ACCUMULATED,
        )

        url = self._report_service.one_time_url(
            buffer, ReportScope.ACCUMULATED.filename(standard, tenant_name)
        )
        response = ReportResponse(tenant, url, fmt=ReportFormat.XLSX)
        return build_response(content=response.dict())

    def _build_questionnaire(
        self,
        collection: ShardsCollection,
        metadata: Metadata,
        cloud: Cloud,
        standard: Standard,
        full_controls: dict[str, int],
        scope: ReportScope,
    ) -> io.BytesIO:
        """Build a questionnaire workbook without constructing an HTTP response."""
        collection.fetch_all()
        rows = self.build_rows(
            collection=collection,
            metadata=metadata,
            cloud=cloud,
            standard=standard,
            full_controls=full_controls,
        )
        buffer = io.BytesIO()
        build_workbook(rows, standard, scope).save(buffer)
        buffer.seek(0)
        return buffer

    # -- data preparation ---------------------------------------------------

    @staticmethod
    def resolve_standard(
        full_cov: dict, standard_name: str
    ) -> Standard | None:
        """Resolves a standard by its full name, i.e. 'CIS Controls v7'."""
        wanted = standard_name.strip().lower()
        for st in full_cov:
            if st.full_name.lower() == wanted or st.name.lower() == wanted:
                return st
        return None

    @staticmethod
    def build_standard_control_to_rule_names(
        metadata: Metadata, cloud: Cloud
    ) -> dict:
        """
        standard name -> version -> control -> {'rules': [...],
        'deprecated': [...], 'severity': Severity}

        Deprecated rules are kept apart: they are excluded from the reported
        rules and from the control severity.
        """
        result = {}
        order = list(Severity)

        def _sev_rank(severity: Severity) -> int:
            if severity is Severity.UNKNOWN:
                return -1
            try:
                return order.index(severity)
            except ValueError:
                return -1

        for rule_name, rule_metadata in metadata.rules.items():
            if rule_metadata.cloud != cloud:
                continue
            deprecated = is_deprecated_rule(rule_name)
            rule_sev = rule_metadata.severity
            for standard, versions in rule_metadata.standard.items():
                std_map = result.setdefault(standard, {})
                for version, controls in versions.items():
                    ver_map = std_map.setdefault(version, {})
                    for control in controls:
                        ctrl_data = ver_map.setdefault(
                            control,
                            {'rules': [], 'deprecated': [], 'severity': None},
                        )
                        if deprecated:
                            ctrl_data['deprecated'].append(rule_name)
                            continue
                        ctrl_data['rules'].append(rule_name)
                        if ctrl_data['severity'] is None or _sev_rank(
                            rule_sev
                        ) > _sev_rank(ctrl_data['severity']):
                            ctrl_data['severity'] = rule_sev
        return result

    def build_rows(
        self,
        collection: ShardsCollection,
        metadata: Metadata,
        cloud: Cloud,
        standard: Standard,
        full_controls: dict[str, int],
        hierarchy: MappingABC[str, dict] | None = None,
    ) -> list[ControlRow]:
        """
        :param hierarchy: optional control id -> {'section_id', 'section_title',
            'sub_sections', 'control_title'} mapping. Once the metadata starts
            providing the standard schema it can be injected here and the
            corresponding columns will appear in the resulting workbook.
        """
        control_to_rules = self.build_standard_control_to_rule_names(
            metadata=metadata, cloud=cloud
        )
        successful_rules, failed_rules = set(), set()
        for part in collection.iter_parts():
            if part.resources:
                failed_rules.add(part.policy)
            else:
                successful_rules.add(part.policy)

        successful = self._report_service.get_standard_to_controls_to_rules(
            it=self._report_service.iter_successful_parts(collection),
            metadata=metadata,
        )
        coverages = calculate_controls_coverages(
            successful.get(standard, {}), full_controls
        )

        aggregated = {control: ControlCompliance() for control in full_controls}
        for control in full_controls:
            data = (
                deep_get(
                    control_to_rules,
                    (standard.name, standard.version_str, control),
                )
                or {}
            )
            control_rules = set(data.get('rules') or ())
            aggregated[control].absorb(
                coverage=coverages.get(control),
                severity=data.get('severity'),
                successful=control_rules & successful_rules,
                failed=control_rules & failed_rules,
                not_evaluated=control_rules - successful_rules - failed_rules,
            )

        rows = []
        for control in sorted(aggregated, key=natural_key):
            compliance = aggregated[control]
            compliance.resolve_overlaps()
            extra = (hierarchy or {}).get(control) or {}
            rows.append(
                ControlRow(
                    control_id=control,
                    section_id=extra.get('section_id', ''),
                    section_title=extra.get('section_title', ''),
                    sub_sections=list(extra.get('sub_sections') or ()),
                    control_title=extra.get('control_title', ''),
                    coverage=compliance.coverage or 0.0,
                    severity=compliance.severity or '',
                    successful_rules=sorted(compliance.successful),
                    failed_rules=sorted(compliance.failed),
                    not_evaluated_rules=sorted(compliance.not_evaluated),
                )
            )
        return rows
