import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from json import dump, load
from pathlib import Path
from re import compile
from typing import Dict

COLUMN_START: str = 'column_start'
COLUMN_END: str = 'column_end'
ROW_START: str = 'row_start'
ROW_END: str = 'row_end'
SUB_ESTIMATION_START: str = 'sub_estimation_start'
SUB_KEY = 'P'
ESTIMATE_KEY = '%'
REQUIREMENT = 'Requirement'
DEFAULT_ESTIMATION_START = 1
ALLOWED_LABEL = r'^[0-9a-zA-Z\.\-\s\(\)]+$'

OUTPUT_KEY = 'output'
COVERAGE_FILENAME = 's3_{cloud}_security_standards_coverage.json'


def _handle_reference(sheet, *data):
    return sheet[';'.join(data)].value


def _handle_string(data: str, sheet):
    _data = tuple()
    for pattern, evaluation in PATTERNS_HANDLER.items():
        _match = compile(pattern).match(data)
        if _match:
            _data = _match.groups()
            break
    else:
        return data
    return evaluation(sheet, *_data)


def _derive_column_index(column: str):
    return ord(column.lower()) % ord('a') + 1


def _pack_estimate(label: str, value: float):
    return {label: {'%': value}}


def _average_out_estimates(subject: dict):
    _key = ESTIMATE_KEY
    _summed = sum((estimate.get(_key, 0) for estimate in subject.values()))
    return _summed / len(subject) if subject else 0


def _derive_aggregation_index(_pending: dict, head: int) -> int:
    """Returns the head key-index, which must commence estimate aggregation."""
    _iterator = (each for each in _pending if each > head and _pending[each])
    return next(_iterator, None)


def _collect_estimate_trail(_pending: dict, dependent: list, head: int,
                            state: dict):
    """
    Failsafe precaution, which mandates an action of collecting any
    non-complemented estimates, which trail ahead of a given head index.
    """
    _trail = dependent[:dependent.index(head) + 1]
    for index, compose in enumerate(_trail[:-1:]):
        refresh = _trail[index + 1]
        if _pending.get(compose):
            _aggregate_estimate(
                _pending=_pending, to_compose=compose,
                to_refresh=refresh, label=state.get(refresh)
            )


def _aggregate_estimate(_pending: dict, to_compose: int, to_refresh: int,
                        label: str):
    """
    Single unit which aggregates estimate collection into a pending one.
    """
    _obj = _pending.setdefault(to_refresh, {})
    _estimate = _pending.pop(to_compose)
    _labeled = _obj.setdefault(label, {})
    _labeled[ESTIMATE_KEY] = _average_out_estimates(_estimate)
    _labeled[SUB_KEY] = _estimate


def _refresh_state(row: list, state: dict, related=True, *subject):
    """
    Meant to yield every index-key, which denotes points at which
    state has gone stale, having refreshed respective keys.
    :returns: Iterator
    """
    for index in subject:
        retained = row[index] if index not in state else state.get(index)
        extracted = retained if row[index] is None and related else row[index]
        condition = (extracted != retained and retained == state.get(index))
        state[index] = extracted
        if condition or not related:
            yield index


def _evaluate_estimation(data, sheet):
    """Evaluation handlers of str and int instances."""
    _value = None
    if isinstance(data, str):
        _value = _handle_string(data, sheet)
    if any(map(lambda _type: isinstance(data, _type), (int, float))):
        _value = data
    return _value


def _i_estimation(row: list, begins: int, state: dict, sheet):
    """
    Yields estimations, based on a range, derived from a given row and
    start index. Under the condition that said range is empty - Points
    are not meant to be composed, so a dummy aggregations are put into place,
    under null keys, which is backward compatible with the standards.
    """
    _default = DEFAULT_ESTIMATION_START
    _subset = row[begins:-1:2]
    _source = _subset or row[_default:-1:2]
    pointer = (begins if _subset else _default) - 2

    for label in _source:
        pointer += 2
        # Manually collects each sub-estimate value and binds it
        # to the key-index(column) -> {$pointer:{$label:{'%':$value}}}
        reference = pointer + 1
        raw = row[reference]
        try:
            _value = float(raw)
        except (ValueError, TypeError):
            # Given the pointer is the default, then self-evaluate
            # Otherwise seek backwards
            reference = pointer - 1 if reference > _default else reference

            _value = _evaluate_estimation(row[reference], sheet)

            _absent = map(lambda each: each is None, (label, _value))

            # Given the value is an unresolvable string or data pair is absent
            if isinstance(_value, str) or all(_absent):
                continue

            # State[parent: $value <- child:$value]
            # Verify that the value has been inherited
            # By checking the state and row labels of the source
            if _value is None and state.get(reference - 1) != row[
                reference - 1]:
                _value = state.setdefault(reference, 0)
            else:
                _value = _value or 0

            _value = float(_value)

        # Seeks label inheritance, given a value has been pointed at such label
        # previously.
        if not label and state.get(pointer + 1) is not None:
            label = state.get(pointer)
        elif not label and pointer + 1 in state:
            continue

        state[reference] = _value

        yield pointer, _pack_estimate(label, _value)


def _parse_sheet(sheet, notation_map: dict):
    _tables = notation_map.get(sheet.title, [])
    _taped_standards = {}

    for t_index, table in enumerate(_tables):
        _point = tuple(f'{table.get(each, "A")}{table.get(ROW_START, 2) - 1}'
                       for each in (COLUMN_START, COLUMN_END))

        _output_format: list = table.get(OUTPUT_KEY) or [sheet.title]
        _output_format += [] if len(_output_format) > 1 else [None]

        _standard_dict = _taped_standards.setdefault(_output_format[0], {})
        _version_dict = _standard_dict.setdefault(_output_format[1], {})

        _header = [each.value for each in sheet[_point[0]:_point[1]][0]]
        _version_dict[_header[0]] = {}

        _sub_estimate_from = table.get(SUB_ESTIMATION_START, 3)

        _table_border = tuple(_derive_column_index(table.get(each))
                              for each in (COLUMN_START, COLUMN_END))
        _row_length = _table_border[1] - _table_border[0]

        # Refreshable string-based data indexes
        _dependent = [0, *range(1, _row_length - 1, 2)][::-1]
        # Meant to store each cell state
        _state: Dict[int] = {}
        # Meant to store pending summed estimate value and a derivable average
        _pending_estimate: Dict[int, Dict[str, float]] = {}
        for r_index, row in enumerate(
                sheet.iter_rows(min_row=table[ROW_START],
                                min_col=_table_border[0],
                                max_row=table[ROW_END],
                                max_col=_table_border[1]),
                start=table[ROW_START]
        ):

            _stale = _state.copy()
            _row = list(map(lambda each: each.value, row))
            # Derives main relation trigger
            step = (_row[0] or _stale.get(0)) == (_stale.get(0) or _row[0])

            # Relation overflow precaution - given an unrecognizable bottom row
            if r_index == table[ROW_END]:
                step = False

            # Yields refreshment key-indexes, which have to be injected as
            # a composed estimation.
            for _refresh in _refresh_state(_row, _state, step, *_dependent):
                _head = _derive_aggregation_index(_pending_estimate, _refresh)
                if _head is not None:
                    _collect_estimate_trail(_pending=_pending_estimate,
                                            dependent=_dependent,
                                            head=_head, state=_stale)
                    _aggregate_estimate(_pending=_pending_estimate,
                                        to_compose=_head, to_refresh=_refresh,
                                        label=_stale.get(_refresh))

            # Derives subsequent labeled estimations, respective state pointers
            for pointer, estimation in _i_estimation(_row, _sub_estimate_from,
                                                     _state, sheet):
                _pending_estimate.setdefault(pointer, {})
                _pending_estimate[pointer].update(estimation)

        coverage = _pending_estimate.get(0, {})
        _version_dict[_header[0]].update(coverage)
        _version_dict[ESTIMATE_KEY] = _average_out_estimates(coverage)
    return _taped_standards


def _parse_sheets(book: 'Workbook', notation: dict):
    _output = {}
    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(_parse_sheet, each, notation): each
                   for each in book.worksheets if each.title in notation}
        for each in as_completed(futures):
            try:
                _result = each.result()
                for standard_key, item in _result.items():
                    _persisted = _output.setdefault(standard_key, item)
                    _persisted.update(item)
                print(f'{futures[each].title} sheet has been parsed.')
            except (TimeoutError, Exception):
                print(f'{futures[each].title} sheet could not be parsed.')
                continue
    return _output


def _retrieve_workbook(file: str):
    from openpyxl import load_workbook
    return load_workbook(filename=file, read_only=True)


def _read_json(notation: str):
    with open(notation, 'r') as fp:
        data = load(fp)
    return data


def _persist_json(parsed: dict, cloud: str):
    folder = '.tmp'
    path = Path.cwd() / folder
    path.mkdir(parents=True, exist_ok=True)
    with open(path / COVERAGE_FILENAME.format(cloud=cloud.lower()), 'w') as fp:
        dump(parsed, fp, separators=(",", ":"))


def parse_standards(file, notation: dict) -> dict:
    return _parse_sheets(_retrieve_workbook(file), notation)


def main(source: str, notation: str, cloud: str):
    book = _retrieve_workbook(source)
    notation = _read_json(notation)
    parsed = _parse_sheets(book, notation)
    _persist_json(parsed, cloud)


PATTERNS_HANDLER = {
    r'^=([A-Z]\d+)$': _handle_reference,
    r'^=AVERAGE\((\w+)(?![\:\;])\)$': _handle_reference
}


def init_parser(parser: argparse.ArgumentParser):
    parser.add_argument('--source', type=str, required=True,
                        help='Path to the xlsx file to parse.')
    parser.add_argument('--notation', type=str, required=True,
                        help='Path to the parsing notation file.')
    parser.add_argument('--cloud', type=lambda x: str(x).lower(),
                        required=True,
                        choices=['aws', 'azure', 'gcp'],
                        help='Cloud of a given security standard.')


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description='Parses security standard xlsx file, exporting to a .json')
    init_parser(parser)
    return parser


if __name__ == '__main__':
    args_parser = build_parser()
    arguments = args_parser.parse_args()
    main(arguments.source, arguments.notation, arguments.cloud)
